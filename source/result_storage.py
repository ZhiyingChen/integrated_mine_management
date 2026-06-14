import logging
import os
import shutil
from typing import Dict

from openpyxl import load_workbook

from .input_data import InputData
from .utils import field, header


class ResultStorage:
    WRITE_DECIMALS = 4

    def __init__(self, input_data: InputData):
        self.input_data = input_data

    def write_to_excel(self):
        raise NotImplementedError("当前阶段不写回 Excel 输出。")

    def write_core_variables_to_excel(
        self,
        sinter_ratio: Dict[int, float],
        pellet_ratio: Dict[int, float],
        burden_ratio: Dict[int, float],
        output_filename: str = None,
        overwrite_source: bool = False,
    ) -> str:
        source_path = self.input_data.excel_path
        if overwrite_source:
            output_path = source_path
        else:
            if output_filename is None:
                raise ValueError("output_filename is required when overwrite_source=False.")
            output_path = os.path.join(self.input_data.exe_folder, field.DATA_DIR, output_filename)
            shutil.copyfile(source_path, output_path)

        workbook = load_workbook(output_path)
        self._write_sheet_values(
            workbook=workbook,
            sheet_name=field.SHEET_INTEGRATED_SINTER,
            values=sinter_ratio,
            col=self.input_data.header_col(field.SHEET_INTEGRATED_SINTER, header.BlendHeader.integrated_ratio),
        )
        self._write_sheet_values(
            workbook=workbook,
            sheet_name=field.SHEET_INTEGRATED_PELLET,
            values=pellet_ratio,
            col=self.input_data.header_col(field.SHEET_INTEGRATED_PELLET, header.BlendHeader.integrated_ratio),
        )
        self._write_sheet_values(
            workbook=workbook,
            sheet_name=field.SHEET_BF_BURDEN,
            values=burden_ratio,
            col=self.input_data.header_col(field.SHEET_BF_BURDEN, header.BurdenHeader.integrated_ratio),
        )
        self._mark_recalculate_on_open(workbook)
        workbook.save(output_path)
        logging.info("core variables written to %s", output_path)
        return output_path

    @staticmethod
    def _write_sheet_values(workbook, sheet_name: str, values: Dict[int, float], col: int):
        sheet = workbook[sheet_name]
        for row, value in values.items():
            sheet.cell(row=row, column=col).value = round(float(value), ResultStorage.WRITE_DECIMALS)

    @staticmethod
    def _mark_recalculate_on_open(workbook):
        calculation = getattr(workbook, "calculation", None)
        if calculation is None:
            return
        if hasattr(calculation, "fullCalcOnLoad"):
            calculation.fullCalcOnLoad = True
        if hasattr(calculation, "forceFullCalc"):
            calculation.forceFullCalc = True
        if hasattr(calculation, "calcMode"):
            calculation.calcMode = "auto"
